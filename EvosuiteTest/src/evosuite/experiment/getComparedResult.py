import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

branch_file_path=".\\branch\\"
fbranch_file_path=".\\fbranch\\"

branch_original_file_names=os.listdir(branch_file_path)
fbranch_original_file_names=os.listdir(fbranch_file_path)

def isNullFile(file):
    wb=load_workbook(file)
    sheets=wb.sheetnames
    if len(sheets)==0:
        return True
    else:
        return False

def removeNullFile(files,path):
    i=0
    while i<len(files):
        if isNullFile(path+files[i]):
            del files[i]
        else:
            i+=1


def removeUnWantedFile(files):
    i=0
    while i<len(files):
        if "30times" in files[i]:
            del files[i]
        else:
            i+=1

removeUnWantedFile(branch_original_file_names)
removeNullFile(branch_original_file_names,branch_file_path)
removeUnWantedFile(fbranch_original_file_names)
removeNullFile(fbranch_original_file_names,fbranch_file_path)

def simplifyFileName(file_path,original_file_names):
    for file in original_file_names:
        target_file=file_path+file
        target_file_name_list=target_file.split("_")
        new_name=target_file_name_list[0]+".xlsx"
        os.rename(target_file,new_name)


#simplify File Name so that openpyxl can work on it
simplifyFileName(branch_file_path,branch_original_file_names)
simplifyFileName(fbranch_file_path,fbranch_original_file_names)

branch_new_file_names=os.listdir(branch_file_path)
fbranch_new_file_names=os.listdir(fbranch_file_path)

removeUnWantedFile(branch_new_file_names)
removeNullFile(branch_new_file_names,branch_file_path)
removeUnWantedFile(fbranch_new_file_names)
removeNullFile(fbranch_new_file_names,fbranch_file_path)

# delete result and create a new one with one 'result' sheet
def checkResultFile():
    files=os.listdir(".\\")
    if "result.xlsx" in files:
        return True
    else:
        return False

if checkResultFile():
    os.remove(".\\result.xlsx")

new_wb=Workbook("result.xlsx")
ws=new_wb.create_sheet('result')
new_wb.save('result.xlsx')


new_wb=load_workbook("result.xlsx")


def checkValidity(ws,k):
    if not isinstance(ws['C'+str(k)].value,float):
        return False
    if not isinstance(ws['D'+str(k)].value,float):
        return False
    if not isinstance(ws['E'+str(k)].value,float):
        return False
    if not isinstance(ws['G'+str(k)].value,float):
        return False
    return True


def count_average(file,ws_title):
    #load file into Workbook
    original_wb=load_workbook(file)
    original_ws=original_wb["data"]
    new_ws=new_wb.create_sheet(title=ws_title)
    #get header
    for column in original_ws['A1':'E1']:
        for cell in column:
            column=get_column_letter(cell.column)
            new_ws[column+str(cell.row)].value=cell.value
    new_ws['F1'].value=original_ws['G1'].value
    #(i,j) represent the same class and the same method
    i=2
    j=3
    # the index in new file
    new_index=2
    while i<=original_ws.max_row:
        #find (i,j) such that (i,j) has the same method
        while j<=original_ws.max_row:
            if original_ws['B'+str(j)].value==original_ws['B'+str(i)].value:
                j=j+1
            else:
                break

        # (i,j-1) contains the same method in the same class
        errornum=0
        total_time=0.0
        total_age=0.0
        total_coverage=0.0
        total_ip_flag_coverage=0.0

        for k in range(i,j):
            if checkValidity(original_ws,k)==True:
                total_time+=original_ws['C'+str(k)].value
                total_coverage+=original_ws['D'+str(k)].value
                total_age+=original_ws['E'+str(k)].value
                total_ip_flag_coverage+=original_ws['G'+str(k)].value
            else:
                errornum+=1

        if errornum!=(j-i+1):
            # count_average
            average_time=total_time/(j-i+1-errornum)
            average_coverage=total_coverage/(j-i+1-errornum)
            average_age=total_age/(j-i+1-errornum)
            average_ip_flag_coverage=total_ip_flag_coverage/(j-i+1-errornum)

            #append new rows
            new_ws['A'+str(new_index)].value=original_ws['A'+str(i)].value
            new_ws['B'+str(new_index)].value=original_ws['B'+str(i)].value
            new_ws['C'+str(new_index)].value=average_time
            new_ws['D'+str(new_index)].value=average_coverage
            new_ws['E'+str(new_index)].value=average_age
            new_ws['F'+str(new_index)].value=average_ip_flag_coverage
        #update parameters
        i=j+1
        j=i+1
        new_index=new_index+1

    new_wb.save("result.xlsx")


files=[]
for file_name in branch_new_file_names:
    if file_name in fbranch_new_file_names:
        if len(file_name.split("_"))==1:
            branch_file=branch_file_path+file_name#.\b\104.xlsx
            branch_ws_name=os.path.splitext(file_name)[0]+"_branch"#101_branch
            fbranch_file=fbranch_file_path+file_name
            fbranch_ws_name=os.path.splitext(file_name)[0]+"_fbranch"
            #generate averge result
            count_average(branch_file,branch_ws_name)
            count_average(fbranch_file,fbranch_ws_name)
            files.append([branch_ws_name,fbranch_ws_name])

#load class and method into list
def load_classes_methods(file_ws):
    classes=[]
    methods=[]
    #load classes
    for row in file_ws['A2':'A'+str(file_ws.max_row)]:
        for cell in row:
            classes.append(cell.value)
    #load methods
    for row in file_ws['B2':'B'+str(file_ws.max_row)]:
        for cell in row:
            methods.append(cell.value)
    return classes,methods

def init_result():
    wb=load_workbook("result.xlsx")
    com_ws=wb["result"]
    #get header
    com_ws['A1'].value="project"
    com_ws['B1'].value="class"
    com_ws['C1'].value="method"
    com_ws['D1'].value="b_execution_time"
    com_ws['E1'].value="b_coverage"
    com_ws['F1'].value="b_age"
    com_ws['G1'].value="b_ip_flag_coverage"
    com_ws['H1'].value="f_execution_time"
    com_ws['I1'].value="f_coverage"
    com_ws['J1'].value="f_age"
    com_ws['K1'].value="f_ip_flag_coverage"
    com_ws['L1'].value="overall"
    wb.save("result.xlsx")


init_result()

def compare(b_ws_title,f_ws_title):
    result_wb=load_workbook("result.xlsx")
    branch_ws=result_wb[b_ws_title]
    fbranch_ws=result_wb[f_ws_title]

    com_ws=result_wb['result']

    b_classes,b_methods=load_classes_methods(branch_ws)
    f_classes,f_methods=load_classes_methods(fbranch_ws)

    new_index=com_ws.max_row+1

    #compare
    for i in range(len(b_methods)):
        if b_methods[i] in f_methods:
            f_index=f_methods.index(b_methods[i])
            if b_classes[i]==f_classes[f_index]:
                if branch_ws['D'+str(i+2)].value<fbranch_ws['D'+str(f_index+2)].value:
                    flag="good coverage"
                elif branch_ws['D'+str(i+2)].value>fbranch_ws['D'+str(f_index+2)].value:
                    flag="bad coverage"
                else:
                    if branch_ws['C'+str(i+2)].value<fbranch_ws['C'+str(f_index+2)].value:
                        flag="bad time"
                    elif branch_ws['C'+str(i+2)].value>fbranch_ws['C'+str(f_index+2)].value:
                        flag="good time"
                    else:
                        flag="tie"
                    if branch_ws['C'+str(i+2)].value>=100 and fbranch_ws['C'+str(f_index+2)].value>=100:
                        flag="tie"


                com_ws['A'+str(new_index)].value=b_ws_title.split("_")[0]
                com_ws['B'+str(new_index)].value=b_classes[i]
                com_ws['C'+str(new_index)].value=b_methods[i]
                com_ws['D'+str(new_index)].value=branch_ws['C'+str(i+2)].value
                com_ws['E'+str(new_index)].value=branch_ws['D'+str(i+2)].value
                com_ws['F'+str(new_index)].value=branch_ws['E'+str(i+2)].value
                com_ws['G'+str(new_index)].value=branch_ws['F'+str(i+2)].value
                com_ws['H'+str(new_index)].value=fbranch_ws['C'+str(f_index+2)].value
                com_ws['I'+str(new_index)].value=fbranch_ws['D'+str(f_index+2)].value
                com_ws['J'+str(new_index)].value=fbranch_ws['E'+str(f_index+2)].value
                com_ws['K'+str(new_index)].value=fbranch_ws['F'+str(f_index+2)].value
                com_ws['L'+str(new_index)].value=flag
                new_index=new_index+1#input to files

    result_wb.save("result.xlsx")

for file in files:
    compare(file[0],file[1])

for file in branch_original_file_names:
    os.rename(branch_file_path+file.split("_")[0]+".xlsx",branch_file_path+file)

for file in fbranch_original_file_names:
    os.rename(fbranch_file_path+file.split("_")[0]+".xlsx",fbranch_file_path+file)
