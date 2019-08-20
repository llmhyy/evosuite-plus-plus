from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


#new_wb=Workbook()
new_wb=load_workbook("result.xlsx")
#new_ws=new_wb.active

def count_average(file_title):
    #load file into Workbook
    original_wb=load_workbook(file_title+".xlsx")
    original_ws=original_wb["data"]
    new_ws=new_wb.create_sheet(title=file_title)
    #get header
    for column in original_ws['A1':'D1']:
        for cell in column:
            column=get_column_letter(cell.column)
            new_ws[column+str(cell.row)].value=cell.value
    #(i,j) represent the same class and the same method
    i=2
    j=3
    # the index in new file
    new_index=2
    while i<=original_ws.max_row:
        #find (i,j) such that (i,j) has the same method
        while j<=original_ws.max_row:
            if original_ws['B'+str(j)].value==original_ws['B'+str(i)].value:
                j=j+1
            else:
                break
        j=j-1

        #average_execution_time
        total_time=0
        for column in original_ws['C'+str(i):'C'+str(j)]:
            for cell in column:
                total_time+=cell.value
        average_time=total_time/(j-i+1)
        #average_coverage
        total_coverage=0
        for column in original_ws['D'+str(i):'D'+str(j)]:
            for cell in column:
                total_coverage+=cell.value
        average_coverage=total_coverage/(j-i+1)
        #append new rows
        new_ws['A'+str(new_index)].value=original_ws['A'+str(i)].value
        new_ws['B'+str(new_index)].value=original_ws['B'+str(i)].value
        new_ws['C'+str(new_index)].value=average_time
        new_ws['D'+str(new_index)].value=average_coverage
        #update parameters
        i=j+1
        j=i+1
        new_index=new_index+1

    new_wb.save("result.xlsx")

files=["101_weka_evotest_branch",
    "101_weka_evotest_fbranch",
    "105_math_evotest_branch",
    "105_math_evotest_fbranch"]

for file in files:
    count_average(file)
