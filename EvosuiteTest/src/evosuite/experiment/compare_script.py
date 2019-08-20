from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

result_wb=load_workbook("result.xlsx")



#load class and method into list
def load_classes_methods(file_ws):
    classes=[]
    methods=[]
    #load classes
    for row in file_ws['A2':'A'+str(file_ws.max_row)]:
        for cell in row:
            classes.append(cell.value)
    #load methods
    for row in file_ws['B2':'B'+str(file_ws.max_row)]:
        for cell in row:
            methods.append(cell.value)
    return classes,methods

def compare(b_ws_title,f_ws_title,ws_title):
    branch_ws=result_wb[b_ws_title]
    fbranch_ws=result_wb[f_ws_title]

    com_ws=result_wb.create_sheet(title=ws_title)

    #get header
    for column in branch_ws['A1':'B1']:
        for cell in column:
            column=get_column_letter(cell.column)
            com_ws[column+str(cell.row)].value=cell.value
    com_ws['C1'].value="result"

    b_classes,b_methods=load_classes_methods(branch_ws)
    f_classes,f_methods=load_classes_methods(fbranch_ws)

    new_index=2

    #compare
    for i in range(len(b_methods)-1):
        if b_methods[i] in f_methods:
            f_index=f_methods.index(b_methods[i])
            if b_classes[i]==f_classes[f_index]:
                if branch_ws['D'+str(i+2)].value>fbranch_ws['D'+str(f_index+2)].value:
                    flag="b"
                elif branch_ws['D'+str(i+2)].value<fbranch_ws['D'+str(f_index+2)].value:
                    flag="f"
                else:
                    if branch_ws['C'+str(i+2)].value>fbranch_ws['C'+str(f_index+2)].value:
                        flag="f"
                    elif branch_ws['C'+str(i+2)].value<fbranch_ws['C'+str(f_index+2)].value:
                        flag="b"
                    else:
                        flag="both"
                com_ws['A'+str(new_index)].value=b_classes[i]
                com_ws['B'+str(new_index)].value=b_methods[i]
                com_ws['C'+str(new_index)].value=flag
                new_index=new_index+1#input to files

    result_wb.save("result.xlsx")

files_title=[["101_weka_evotest_branch","101_weka_evotest_fbranch","101_weka_evotest"],
            ["105_math_evotest_branch","105_math_evotest_fbranch","105_math_evotest"]]

for ele in files_title:
    compare(ele[0],ele[1],ele[2])
